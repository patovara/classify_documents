"""
=============================================================================
SISTEMA DE COTIZACIÓN EMPRESARIAL REMAA
=============================================================================

PROPÓSITO:
Aplicación de uso diario para crear cotizaciones rápidamente reutilizando
una base centralizada de conceptos validados.

VERSIÓN: 2.0
EMPRESA: REMAA
AUTOR: Sistema de Automatización Empresarial
FECHA: 2026-02-03

DESCRIPCIÓN:
Interfaz gráfica intuitiva que permite:
✅ Búsqueda rápida de conceptos por clave o descripción
✅ Agregar conceptos a cotización con doble clic
✅ Ajustar cantidades y márgenes de ganancia
✅ Cálculo automático de totales
✅ Prevención de duplicados
✅ Exportación profesional a Excel

REQUISITOS PREVIOS:
1. Instalar dependencias:
   pip install pandas openpyxl

2. Tener el archivo:
   conceptos_master.csv
   
   Con las columnas:
   - clave
   - concepto_original
   - unidad
   - precio_unitario
   - origen_pdf

EJECUCIÓN:
   python sistema_cotizacion_remaa.py

INSTRUCCIONES DE USO:
1. Iniciar la aplicación
2. Buscar conceptos por palabra clave en el campo de búsqueda
3. Hacer DOBLE CLIC en un concepto para agregarlo a la cotización
4. Especificar cantidad y margen de ganancia (%)
5. Ajustar conceptos si es necesario
6. Click en "EXPORTAR COTIZACIÓN"
7. Ingresar datos del proyecto y guardar

CARACTERÍSTICAS:
✅ Búsqueda instantánea e insensible a mayúsculas
✅ No permite duplicados en la misma cotización
✅ Cálculo automático de precios finales y subtotales
✅ Exportación profesional a Excel con formato
✅ Validación de datos antes de exportar
✅ Interfaz limpia y profesional

RESTRICCIONES:
- NO modifica conceptos_master.csv
- NO lee PDFs
- NO usa machine learning
- Prioriza rapidez y facilidad de uso

=============================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime
import os
from pathlib import Path
from typing import Optional, List, Dict


# ============================================================================
# CLASE PRINCIPAL
# ============================================================================

class SistemaCotizacionREMAA:
    """
    Sistema completo de gestión de cotizaciones para REMAA.
    Optimizado para uso diario con máxima eficiencia operativa.
    """
    
    def __init__(self, root: tk.Tk):
        """
        Inicializa la aplicación y sus componentes.
        
        Args:
            root: Ventana principal de Tkinter
        """
        self.root = root
        self.root.title("Sistema de Cotización REMAA")
        self.root.geometry("1300x750")
        self.root.configure(bg="#f0f0f0")
        
        # Variables de estado
        self.conceptos_df: Optional[pd.DataFrame] = None
        self.cotizacion_actual: List[Dict] = []
        self.archivo_master = "conceptos_master.csv"
        self.carpeta_exportaciones = "cotizaciones_remaa"
        
        # Crear carpeta de exportaciones
        Path(self.carpeta_exportaciones).mkdir(exist_ok=True)
        
        # Cargar base de conceptos
        if not self.cargar_conceptos_master():
            return
        
        # Construir interfaz
        self.construir_interfaz()
        
        # Cargar todos los conceptos inicialmente
        self.actualizar_tabla_conceptos()
    
    
    def cargar_conceptos_master(self) -> bool:
        """
        Carga el archivo maestro de conceptos.
        
        Returns:
            bool: True si se cargó correctamente
        """
        if not os.path.exists(self.archivo_master):
            messagebox.showerror(
                "❌ Archivo no encontrado",
                f"No se encontró el archivo: {self.archivo_master}\n\n"
                "Este archivo debe ser generado previamente por el\n"
                "MÓDULO 1 - Extractor de Conceptos.\n\n"
                "Asegúrese de que el archivo esté en el mismo directorio."
            )
            self.root.destroy()
            return False
        
        try:
            # Cargar CSV
            self.conceptos_df = pd.read_csv(self.archivo_master, encoding='utf-8-sig')
            
            # Validar columnas requeridas
            columnas_requeridas = ['clave', 'concepto_original', 'unidad', 'precio_unitario']
            columnas_faltantes = [col for col in columnas_requeridas 
                                 if col not in self.conceptos_df.columns]
            
            if columnas_faltantes:
                messagebox.showerror(
                    "❌ Formato Inválido",
                    f"El archivo {self.archivo_master} no contiene las columnas requeridas:\n"
                    f"{', '.join(columnas_faltantes)}"
                )
                self.root.destroy()
                return False
            
            # Limpiar datos
            self.conceptos_df['clave'] = self.conceptos_df['clave'].astype(str).str.strip()
            self.conceptos_df['concepto_original'] = self.conceptos_df['concepto_original'].astype(str).str.strip()
            self.conceptos_df['unidad'] = self.conceptos_df['unidad'].astype(str).str.strip()
            self.conceptos_df['precio_unitario'] = pd.to_numeric(
                self.conceptos_df['precio_unitario'], errors='coerce'
            ).fillna(0)
            
            # Eliminar filas sin clave
            self.conceptos_df = self.conceptos_df[
                self.conceptos_df['clave'].notna() & 
                (self.conceptos_df['clave'] != '') &
                (self.conceptos_df['clave'] != 'nan')
            ]
            
            messagebox.showinfo(
                "✅ Base de Datos Cargada",
                f"Se cargaron correctamente:\n\n"
                f"📊 {len(self.conceptos_df)} conceptos únicos\n"
                f"📁 Fuente: {self.archivo_master}"
            )
            
            return True
            
        except Exception as e:
            messagebox.showerror(
                "❌ Error al Cargar",
                f"Error al cargar {self.archivo_master}:\n\n{str(e)}"
            )
            self.root.destroy()
            return False
    
    
    def construir_interfaz(self):
        """Construye la interfaz gráfica completa."""
        
        # ==================== BARRA DE TÍTULO ====================
        frame_titulo = tk.Frame(self.root, bg="#1a237e", height=70)
        frame_titulo.pack(fill=tk.X)
        frame_titulo.pack_propagate(False)
        
        tk.Label(
            frame_titulo,
            text="🏢 SISTEMA DE COTIZACIÓN REMAA",
            font=("Segoe UI", 20, "bold"),
            bg="#1a237e",
            fg="white"
        ).pack(expand=True)
        
        # ==================== FRAME DE BÚSQUEDA ====================
        frame_busqueda = tk.LabelFrame(
            self.root,
            text=" 🔍 BÚSQUEDA DE CONCEPTOS ",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            padx=20,
            pady=15
        )
        frame_busqueda.pack(fill=tk.X, padx=15, pady=15)
        
        frame_busqueda_content = tk.Frame(frame_busqueda, bg="#ffffff")
        frame_busqueda_content.pack(fill=tk.X)
        
        tk.Label(
            frame_busqueda_content,
            text="Buscar por CLAVE o CONCEPTO:",
            font=("Segoe UI", 10, "bold"),
            bg="#ffffff"
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        self.entrada_busqueda = tk.Entry(
            frame_busqueda_content,
            width=60,
            font=("Segoe UI", 11)
        )
        self.entrada_busqueda.pack(side=tk.LEFT, padx=5)
        self.entrada_busqueda.bind("<KeyRelease>", lambda e: self.buscar_conceptos())
        
        tk.Button(
            frame_busqueda_content,
            text="🔍 Buscar",
            command=self.buscar_conceptos,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=20,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_busqueda_content,
            text="↻ Limpiar",
            command=self.limpiar_busqueda,
            bg="#FF9800",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=20,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        self.label_resultados = tk.Label(
            frame_busqueda_content,
            text="",
            font=("Segoe UI", 9, "italic"),
            bg="#ffffff",
            fg="#666666"
        )
        self.label_resultados.pack(side=tk.LEFT, padx=20)
        
        # ==================== TABLA DE CONCEPTOS DISPONIBLES ====================
        frame_conceptos = tk.LabelFrame(
            self.root,
            text=" 📋 CONCEPTOS DISPONIBLES (Doble clic para agregar) ",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            padx=10,
            pady=10
        )
        frame_conceptos.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        # Container con scrollbars
        container_conceptos = tk.Frame(frame_conceptos, bg="#ffffff")
        container_conceptos.pack(fill=tk.BOTH, expand=True)
        
        scroll_y_conceptos = tk.Scrollbar(container_conceptos, orient=tk.VERTICAL)
        scroll_x_conceptos = tk.Scrollbar(container_conceptos, orient=tk.HORIZONTAL)
        
        # Configurar estilo de la tabla
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background="#ffffff",
            foreground="#000000",
            rowheight=30,
            fieldbackground="#ffffff",
            font=("Segoe UI", 10)
        )
        style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#1a237e",
            foreground="white"
        )
        style.map('Treeview', background=[('selected', '#2196F3')])
        
        self.tabla_conceptos = ttk.Treeview(
            container_conceptos,
            columns=("CLAVE", "CONCEPTO", "UNIDAD", "PRECIO_BASE"),
            show="headings",
            yscrollcommand=scroll_y_conceptos.set,
            xscrollcommand=scroll_x_conceptos.set,
            height=6
        )
        
        scroll_y_conceptos.config(command=self.tabla_conceptos.yview)
        scroll_x_conceptos.config(command=self.tabla_conceptos.xview)
        
        # Configurar columnas
        self.tabla_conceptos.heading("CLAVE", text="CLAVE")
        self.tabla_conceptos.heading("CONCEPTO", text="CONCEPTO")
        self.tabla_conceptos.heading("UNIDAD", text="UNIDAD")
        self.tabla_conceptos.heading("PRECIO_BASE", text="PRECIO BASE")
        
        self.tabla_conceptos.column("CLAVE", width=100, anchor=tk.CENTER)
        self.tabla_conceptos.column("CONCEPTO", width=800, anchor=tk.W)
        self.tabla_conceptos.column("UNIDAD", width=100, anchor=tk.CENTER)
        self.tabla_conceptos.column("PRECIO_BASE", width=120, anchor=tk.E)
        
        # Evento doble clic
        self.tabla_conceptos.bind("<Double-1>", self.agregar_a_cotizacion)
        
        self.tabla_conceptos.grid(row=0, column=0, sticky="nsew")
        scroll_y_conceptos.grid(row=0, column=1, sticky="ns")
        scroll_x_conceptos.grid(row=1, column=0, sticky="ew")
        
        container_conceptos.grid_rowconfigure(0, weight=1)
        container_conceptos.grid_columnconfigure(0, weight=1)
        
        # ==================== COTIZACIÓN ACTUAL ====================
        frame_cotizacion = tk.LabelFrame(
            self.root,
            text=" 📝 COTIZACIÓN ACTUAL ",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            padx=10,
            pady=10
        )
        frame_cotizacion.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        container_cotizacion = tk.Frame(frame_cotizacion, bg="#ffffff")
        container_cotizacion.pack(fill=tk.BOTH, expand=True)
        
        scroll_y_cot = tk.Scrollbar(container_cotizacion, orient=tk.VERTICAL)
        scroll_x_cot = tk.Scrollbar(container_cotizacion, orient=tk.HORIZONTAL)
        
        self.tabla_cotizacion = ttk.Treeview(
            container_cotizacion,
            columns=("CLAVE", "CONCEPTO", "CANTIDAD", "UNIDAD", "P_BASE", "MARGEN", "P_FINAL", "SUBTOTAL"),
            show="headings",
            yscrollcommand=scroll_y_cot.set,
            xscrollcommand=scroll_x_cot.set,
            height=5
        )
        
        scroll_y_cot.config(command=self.tabla_cotizacion.yview)
        scroll_x_cot.config(command=self.tabla_cotizacion.xview)
        
        self.tabla_cotizacion.heading("CLAVE", text="CLAVE")
        self.tabla_cotizacion.heading("CONCEPTO", text="CONCEPTO")
        self.tabla_cotizacion.heading("CANTIDAD", text="CANTIDAD")
        self.tabla_cotizacion.heading("UNIDAD", text="UNIDAD")
        self.tabla_cotizacion.heading("P_BASE", text="P. BASE")
        self.tabla_cotizacion.heading("MARGEN", text="MARGEN %")
        self.tabla_cotizacion.heading("P_FINAL", text="P. FINAL")
        self.tabla_cotizacion.heading("SUBTOTAL", text="SUBTOTAL")
        
        self.tabla_cotizacion.column("CLAVE", width=90, anchor=tk.CENTER)
        self.tabla_cotizacion.column("CONCEPTO", width=500, anchor=tk.W)
        self.tabla_cotizacion.column("CANTIDAD", width=90, anchor=tk.CENTER)
        self.tabla_cotizacion.column("UNIDAD", width=80, anchor=tk.CENTER)
        self.tabla_cotizacion.column("P_BASE", width=100, anchor=tk.E)
        self.tabla_cotizacion.column("MARGEN", width=90, anchor=tk.CENTER)
        self.tabla_cotizacion.column("P_FINAL", width=100, anchor=tk.E)
        self.tabla_cotizacion.column("SUBTOTAL", width=120, anchor=tk.E)
        
        self.tabla_cotizacion.grid(row=0, column=0, sticky="nsew")
        scroll_y_cot.grid(row=0, column=1, sticky="ns")
        scroll_x_cot.grid(row=1, column=0, sticky="ew")
        
        container_cotizacion.grid_rowconfigure(0, weight=1)
        container_cotizacion.grid_columnconfigure(0, weight=1)
        
        # ==================== PANEL DE ACCIONES ====================
        frame_acciones = tk.Frame(self.root, bg="#263238", padx=20, pady=20)
        frame_acciones.pack(fill=tk.X, padx=15, pady=(0, 10), side=tk.BOTTOM)
        
        # Botones de acción
        tk.Button(
            frame_acciones,
            text="✏️ Editar Cantidad/Margen",
            command=self.editar_concepto_cotizacion,
            bg="#FFC107",
            fg="#000000",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_acciones,
            text="❌ Eliminar Seleccionado",
            command=self.eliminar_de_cotizacion,
            bg="#F44336",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_acciones,
            text="🗑️ Limpiar Todo",
            command=self.limpiar_cotizacion,
            bg="#E91E63",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=8,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        # Total
        self.label_total = tk.Label(
            frame_acciones,
            text="TOTAL: $0.00",
            font=("Segoe UI", 18, "bold"),
            bg="#263238",
            fg="#4CAF50"
        )
        self.label_total.pack(side=tk.LEFT, padx=40)
        
        # Botón exportar
        tk.Button(
            frame_acciones,
            text="💾 EXPORTAR COTIZACIÓN",
            command=self.exportar_cotizacion,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            padx=30,
            pady=12,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.RIGHT, padx=5)
    
    
    def actualizar_tabla_conceptos(self, df_filtrado: Optional[pd.DataFrame] = None):
        """
        Actualiza la tabla de conceptos disponibles.
        
        Args:
            df_filtrado: DataFrame filtrado o None para mostrar todos
        """
        # Limpiar tabla
        for item in self.tabla_conceptos.get_children():
            self.tabla_conceptos.delete(item)
        
        # DataFrame a mostrar
        df = df_filtrado if df_filtrado is not None else self.conceptos_df
        
        # Insertar filas
        for _, row in df.iterrows():
            self.tabla_conceptos.insert("", tk.END, values=(
                row['clave'],
                row['concepto_original'][:100] + "..." if len(row['concepto_original']) > 100 
                    else row['concepto_original'],
                row['unidad'],
                f"${row['precio_unitario']:,.2f}"
            ))
        
        # Actualizar contador
        self.label_resultados.config(text=f"Mostrando {len(df)} conceptos")
    
    
    def buscar_conceptos(self):
        """Busca conceptos por texto en clave o concepto."""
        texto = self.entrada_busqueda.get().strip().lower()
        
        if not texto:
            self.actualizar_tabla_conceptos()
            return
        
        # Filtrar por clave o concepto
        df_filtrado = self.conceptos_df[
            self.conceptos_df['clave'].str.lower().str.contains(texto, na=False) |
            self.conceptos_df['concepto_original'].str.lower().str.contains(texto, na=False)
        ]
        
        self.actualizar_tabla_conceptos(df_filtrado)
    
    
    def limpiar_busqueda(self):
        """Limpia el campo de búsqueda y muestra todos los conceptos."""
        self.entrada_busqueda.delete(0, tk.END)
        self.actualizar_tabla_conceptos()
    
    
    def clave_existe_en_cotizacion(self, clave: str) -> bool:
        """
        Verifica si una clave ya existe en la cotización actual.
        
        Args:
            clave: Clave a verificar
            
        Returns:
            bool: True si ya existe
        """
        return any(item['clave'] == clave for item in self.cotizacion_actual)
    
    
    def agregar_a_cotizacion(self, event):
        """Abre diálogo para agregar concepto a la cotización."""
        seleccion = self.tabla_conceptos.selection()
        if not seleccion:
            return
        
        # Obtener concepto seleccionado
        item = self.tabla_conceptos.item(seleccion[0])
        clave = item['values'][0]
        
        # Verificar duplicados
        if self.clave_existe_en_cotizacion(clave):
            messagebox.showwarning(
                "⚠️ Concepto Duplicado",
                f"La clave '{clave}' ya está en la cotización actual.\n\n"
                "No se permiten conceptos duplicados.\n"
                "Si desea modificar cantidad o margen, use el botón 'Editar'."
            )
            return
        
        # Obtener datos completos del concepto
        concepto_df = self.conceptos_df[self.conceptos_df['clave'] == clave].iloc[0]
        
        # Ventana de diálogo
        dialogo = tk.Toplevel(self.root)
        dialogo.title("Agregar Concepto a Cotización")
        dialogo.geometry("600x400")
        dialogo.transient(self.root)
        dialogo.grab_set()
        dialogo.configure(bg="#f5f5f5")
        
        # Título
        tk.Label(
            dialogo,
            text="➕ AGREGAR CONCEPTO",
            font=("Segoe UI", 14, "bold"),
            bg="#1a237e",
            fg="white",
            pady=15
        ).pack(fill=tk.X)
        
        # Información del concepto
        frame_info = tk.Frame(dialogo, bg="#f5f5f5", padx=20, pady=20)
        frame_info.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(
            frame_info,
            text=f"Clave: {concepto_df['clave']}",
            font=("Segoe UI", 10, "bold"),
            bg="#f5f5f5",
            anchor=tk.W
        ).pack(fill=tk.X, pady=5)
        
        tk.Label(
            frame_info,
            text=f"Concepto:",
            font=("Segoe UI", 10, "bold"),
            bg="#f5f5f5",
            anchor=tk.W
        ).pack(fill=tk.X)
        
        text_concepto = tk.Text(
            frame_info,
            height=4,
            font=("Segoe UI", 9),
            wrap=tk.WORD,
            bg="#ffffff",
            relief=tk.SOLID,
            borderwidth=1
        )
        text_concepto.pack(fill=tk.X, pady=5)
        text_concepto.insert("1.0", concepto_df['concepto_original'])
        text_concepto.config(state=tk.DISABLED)
        
        tk.Label(
            frame_info,
            text=f"Unidad: {concepto_df['unidad']}   |   Precio Base: ${concepto_df['precio_unitario']:,.2f}",
            font=("Segoe UI", 10),
            bg="#f5f5f5",
            anchor=tk.W
        ).pack(fill=tk.X, pady=10)
        
        # Separador
        ttk.Separator(frame_info, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        
        # Campos de entrada
        frame_campos = tk.Frame(frame_info, bg="#f5f5f5")
        frame_campos.pack(fill=tk.X, pady=10)
        
        tk.Label(
            frame_campos,
            text="Cantidad:",
            font=("Segoe UI", 11, "bold"),
            bg="#f5f5f5"
        ).grid(row=0, column=0, sticky=tk.E, padx=10, pady=8)
        
        entrada_cantidad = tk.Entry(frame_campos, width=15, font=("Segoe UI", 11))
        entrada_cantidad.grid(row=0, column=1, padx=10, pady=8)
        entrada_cantidad.insert(0, "1")
        entrada_cantidad.focus()
        
        tk.Label(
            frame_campos,
            text="Margen de Ganancia (%):",
            font=("Segoe UI", 11, "bold"),
            bg="#f5f5f5"
        ).grid(row=1, column=0, sticky=tk.E, padx=10, pady=8)
        
        entrada_margen = tk.Entry(frame_campos, width=15, font=("Segoe UI", 11))
        entrada_margen.grid(row=1, column=1, padx=10, pady=8)
        entrada_margen.insert(0, "0")
        
        # Label para mostrar cálculo
        label_calculo = tk.Label(
            frame_info,
            text="",
            font=("Segoe UI", 10),
            bg="#f5f5f5",
            fg="#2196F3"
        )
        label_calculo.pack(pady=10)
        
        def actualizar_calculo(*args):
            """Actualiza el cálculo en tiempo real."""
            try:
                cantidad = float(entrada_cantidad.get() or 0)
                margen = float(entrada_margen.get() or 0)
                precio_base = concepto_df['precio_unitario']
                precio_final = precio_base * (1 + margen / 100)
                subtotal = cantidad * precio_final
                
                label_calculo.config(
                    text=f"💰 Precio Final: ${precio_final:,.2f}   |   "
                         f"📊 Subtotal: ${subtotal:,.2f}"
                )
            except:
                label_calculo.config(text="")
        
        entrada_cantidad.bind("<KeyRelease>", actualizar_calculo)
        entrada_margen.bind("<KeyRelease>", actualizar_calculo)
        actualizar_calculo()
        
        def confirmar():
            """Confirma y agrega el concepto a la cotización."""
            try:
                cantidad = float(entrada_cantidad.get())
                margen = float(entrada_margen.get())
                
                if cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                    return
                
                precio_base = concepto_df['precio_unitario']
                precio_final = precio_base * (1 + margen / 100)
                subtotal = cantidad * precio_final
                
                # Agregar a cotización
                self.cotizacion_actual.append({
                    'clave': concepto_df['clave'],
                    'concepto': concepto_df['concepto_original'],
                    'cantidad': cantidad,
                    'unidad': concepto_df['unidad'],
                    'precio_base': precio_base,
                    'margen_porcentaje': margen,
                    'precio_final': precio_final,
                    'subtotal': subtotal
                })
                
                self.actualizar_tabla_cotizacion()
                dialogo.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "Ingrese valores numéricos válidos")
        
        # Botones
        frame_botones = tk.Frame(dialogo, bg="#f5f5f5", pady=15)
        frame_botones.pack(fill=tk.X)
        
        tk.Button(
            frame_botones,
            text="✓ Agregar a Cotización",
            command=confirmar,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=20,
            pady=10,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=20)
        
        tk.Button(
            frame_botones,
            text="✗ Cancelar",
            command=dialogo.destroy,
            bg="#9E9E9E",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=20,
            pady=10,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        dialogo.bind("<Return>", lambda e: confirmar())
        dialogo.bind("<Escape>", lambda e: dialogo.destroy())
    
    
    def actualizar_tabla_cotizacion(self):
        """Actualiza la tabla de cotización y el total general."""
        # Limpiar tabla
        for item in self.tabla_cotizacion.get_children():
            self.tabla_cotizacion.delete(item)
        
        # Insertar conceptos
        total_general = 0
        for item in self.cotizacion_actual:
            self.tabla_cotizacion.insert("", tk.END, values=(
                item['clave'],
                item['concepto'][:80] + "..." if len(item['concepto']) > 80 else item['concepto'],
                f"{item['cantidad']:.2f}",
                item['unidad'],
                f"${item['precio_base']:,.2f}",
                f"{item['margen_porcentaje']:.1f}%",
                f"${item['precio_final']:,.2f}",
                f"${item['subtotal']:,.2f}"
            ))
            total_general += item['subtotal']
        
        # Actualizar total
        self.label_total.config(text=f"TOTAL: ${total_general:,.2f}")
    
    
    def editar_concepto_cotizacion(self):
        """Permite editar cantidad y margen de un concepto en la cotización."""
        seleccion = self.tabla_cotizacion.selection()
        if not seleccion:
            messagebox.showwarning(
                "⚠️ Selección requerida",
                "Seleccione un concepto de la cotización para editar."
            )
            return
        
        # Obtener índice del concepto
        index = self.tabla_cotizacion.index(seleccion[0])
        concepto = self.cotizacion_actual[index]
        
        # Ventana de edición
        dialogo = tk.Toplevel(self.root)
        dialogo.title("Editar Concepto")
        dialogo.geometry("500x300")
        dialogo.transient(self.root)
        dialogo.grab_set()
        dialogo.configure(bg="#f5f5f5")
        
        tk.Label(
            dialogo,
            text="✏️ EDITAR CONCEPTO",
            font=("Segoe UI", 14, "bold"),
            bg="#1a237e",
            fg="white",
            pady=15
        ).pack(fill=tk.X)
        
        frame = tk.Frame(dialogo, bg="#f5f5f5", padx=30, pady=30)
        frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(
            frame,
            text=f"Clave: {concepto['clave']}",
            font=("Segoe UI", 11, "bold"),
            bg="#f5f5f5"
        ).pack(anchor=tk.W, pady=5)
        
        tk.Label(
            frame,
            text="Nueva Cantidad:",
            font=("Segoe UI", 10, "bold"),
            bg="#f5f5f5"
        ).pack(anchor=tk.W, pady=(15, 5))
        
        entrada_cantidad = tk.Entry(frame, width=20, font=("Segoe UI", 11))
        entrada_cantidad.pack(anchor=tk.W)
        entrada_cantidad.insert(0, str(concepto['cantidad']))
        entrada_cantidad.focus()
        
        tk.Label(
            frame,
            text="Nuevo Margen (%):",
            font=("Segoe UI", 10, "bold"),
            bg="#f5f5f5"
        ).pack(anchor=tk.W, pady=(15, 5))
        
        entrada_margen = tk.Entry(frame, width=20, font=("Segoe UI", 11))
        entrada_margen.pack(anchor=tk.W)
        entrada_margen.insert(0, str(concepto['margen_porcentaje']))
        
        def confirmar_edicion():
            try:
                nueva_cantidad = float(entrada_cantidad.get())
                nuevo_margen = float(entrada_margen.get())
                
                if nueva_cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                    return
                
                # Actualizar concepto
                precio_final = concepto['precio_base'] * (1 + nuevo_margen / 100)
                concepto['cantidad'] = nueva_cantidad
                concepto['margen_porcentaje'] = nuevo_margen
                concepto['precio_final'] = precio_final
                concepto['subtotal'] = nueva_cantidad * precio_final
                
                self.actualizar_tabla_cotizacion()
                dialogo.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "Ingrese valores numéricos válidos")
        
        frame_btn = tk.Frame(dialogo, bg="#f5f5f5", pady=15)
        frame_btn.pack()
        
        tk.Button(
            frame_btn,
            text="✓ Guardar Cambios",
            command=confirmar_edicion,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=20,
            pady=8,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_btn,
            text="✗ Cancelar",
            command=dialogo.destroy,
            bg="#9E9E9E",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=20,
            pady=8,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        dialogo.bind("<Return>", lambda e: confirmar_edicion())
        dialogo.bind("<Escape>", lambda e: dialogo.destroy())
    
    
    def eliminar_de_cotizacion(self):
        """Elimina el concepto seleccionado de la cotización."""
        seleccion = self.tabla_cotizacion.selection()
        if not seleccion:
            messagebox.showwarning(
                "⚠️ Selección requerida",
                "Seleccione un concepto para eliminar."
            )
            return
        
        # Confirmar eliminación
        if messagebox.askyesno(
            "Confirmar eliminación",
            "¿Está seguro de eliminar este concepto de la cotización?"
        ):
            index = self.tabla_cotizacion.index(seleccion[0])
            del self.cotizacion_actual[index]
            self.actualizar_tabla_cotizacion()
    
    
    def limpiar_cotizacion(self):
        """Limpia toda la cotización actual."""
        if not self.cotizacion_actual:
            return
        
        if messagebox.askyesno(
            "⚠️ Confirmar limpieza",
            "¿Está seguro de eliminar todos los conceptos de la cotización actual?\n\n"
            "Esta acción no se puede deshacer."
        ):
            self.cotizacion_actual.clear()
            self.actualizar_tabla_cotizacion()
    
    
    def exportar_cotizacion(self):
        """Exporta la cotización a Excel."""
        if not self.cotizacion_actual:
            messagebox.showwarning(
                "⚠️ Cotización Vacía",
                "Agregue al menos un concepto antes de exportar."
            )
            return
        
        # Diálogo para datos del proyecto
        dialogo = tk.Toplevel(self.root)
        dialogo.title("Datos del Proyecto")
        dialogo.geometry("550x350")
        dialogo.transient(self.root)
        dialogo.grab_set()
        dialogo.configure(bg="#f5f5f5")
        
        tk.Label(
            dialogo,
            text="📋 INFORMACIÓN DEL PROYECTO",
            font=("Segoe UI", 14, "bold"),
            bg="#1a237e",
            fg="white",
            pady=15
        ).pack(fill=tk.X)
        
        frame = tk.Frame(dialogo, bg="#f5f5f5", padx=30, pady=30)
        frame.pack(fill=tk.BOTH, expand=True)
        
        campos = []
        labels = [
            ("ID del Proyecto:", True),
            ("Nombre del Cliente:", False),
            ("Contacto:", False),
            ("Observaciones:", False)
        ]
        
        for label_text, requerido in labels:
            tk.Label(
                frame,
                text=label_text + (" *" if requerido else ""),
                font=("Segoe UI", 10, "bold"),
                bg="#f5f5f5",
                fg="#d32f2f" if requerido else "#000000"
            ).pack(anchor=tk.W, pady=(10, 5))
            
            entrada = tk.Entry(frame, width=40, font=("Segoe UI", 10))
            entrada.pack(anchor=tk.W)
            campos.append(entrada)
        
        campos[0].focus()
        
        def confirmar_exportacion():
            id_proyecto = campos[0].get().strip()
            if not id_proyecto:
                messagebox.showerror("Error", "El ID del proyecto es obligatorio")
                return
            
            cliente = campos[1].get().strip() or "N/A"
            contacto = campos[2].get().strip() or "N/A"
            observaciones = campos[3].get().strip() or ""
            
            dialogo.destroy()
            self.generar_excel(id_proyecto, cliente, contacto, observaciones)
        
        frame_btn = tk.Frame(dialogo, bg="#f5f5f5", pady=15)
        frame_btn.pack()
        
        tk.Button(
            frame_btn,
            text="💾 Exportar a Excel",
            command=confirmar_exportacion,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=25,
            pady=10,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_btn,
            text="✗ Cancelar",
            command=dialogo.destroy,
            bg="#9E9E9E",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=25,
            pady=10,
            cursor="hand2",
            relief=tk.FLAT
        ).pack(side=tk.LEFT, padx=5)
        
        dialogo.bind("<Return>", lambda e: confirmar_exportacion())
        dialogo.bind("<Escape>", lambda e: dialogo.destroy())
    
    
    def generar_excel(self, id_proyecto: str, cliente: str, 
                     contacto: str, observaciones: str):
        """
        Genera el archivo Excel con la cotización.
        
        Args:
            id_proyecto: ID del proyecto
            cliente: Nombre del cliente
            contacto: Persona de contacto
            observaciones: Observaciones adicionales
        """
        # Generar nombre de archivo
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"COT_REMAA_{id_proyecto.replace(' ', '_')}_{fecha}.xlsx"
        
        # Solicitar ubicación de guardado
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_archivo,
            initialdir=self.carpeta_exportaciones
        )
        
        if not ruta_archivo:
            return
        
        try:
            # Crear DataFrame de la cotización
            df_cotizacion = pd.DataFrame(self.cotizacion_actual)
            
            # Calcular total
            total = df_cotizacion['subtotal'].sum()
            
            # Crear Excel
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                # Hoja 1: Información del proyecto
                df_info = pd.DataFrame({
                    'Campo': [
                        'ID Proyecto',
                        'Cliente',
                        'Contacto',
                        'Fecha de Cotización',
                        'Observaciones',
                        '',
                        'TOTAL GENERAL'
                    ],
                    'Valor': [
                        id_proyecto,
                        cliente,
                        contacto,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        observaciones,
                        '',
                        f"${total:,.2f}"
                    ]
                })
                df_info.to_excel(writer, sheet_name='Información', index=False)
                
                # Hoja 2: Cotización detallada
                df_export = df_cotizacion[[
                    'clave', 'concepto', 'cantidad', 'unidad',
                    'precio_base', 'margen_porcentaje', 'precio_final', 'subtotal'
                ]].copy()
                
                df_export.columns = [
                    'CLAVE', 'CONCEPTO', 'CANTIDAD', 'UNIDAD',
                    'PRECIO BASE', 'MARGEN %', 'PRECIO FINAL', 'SUBTOTAL'
                ]
                
                df_export.to_excel(writer, sheet_name='Cotización', index=False)
                
                # Aplicar formato profesional
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                
                # Formatear hoja de información
                ws_info = writer.sheets['Información']
                ws_info.column_dimensions['A'].width = 25
                ws_info.column_dimensions['B'].width = 50
                
                for cell in ws_info[1]:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a237e", 
                                           end_color="1a237e", fill_type="solid")
                
                # Resaltar total
                ws_info[f'A{len(df_info) + 1}'].font = Font(bold=True, size=14)
                ws_info[f'B{len(df_info) + 1}'].font = Font(bold=True, size=14, color="2e7d32")
                
                # Formatear hoja de cotización
                ws_cot = writer.sheets['Cotización']
                
                # Encabezados
                for cell in ws_cot[1]:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a237e", 
                                           end_color="1a237e", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar anchos de columna
                ws_cot.column_dimensions['A'].width = 12
                ws_cot.column_dimensions['B'].width = 60
                ws_cot.column_dimensions['C'].width = 12
                ws_cot.column_dimensions['D'].width = 10
                ws_cot.column_dimensions['E'].width = 15
                ws_cot.column_dimensions['F'].width = 12
                ws_cot.column_dimensions['G'].width = 15
                ws_cot.column_dimensions['H'].width = 15
                
                # Agregar fila de total
                ultima_fila = len(df_export) + 2
                ws_cot.cell(row=ultima_fila + 1, column=7, value="TOTAL:")
                ws_cot.cell(row=ultima_fila + 1, column=8, value=total)
                
                # Formato de total
                ws_cot.cell(row=ultima_fila + 1, column=7).font = Font(bold=True, size=12)
                ws_cot.cell(row=ultima_fila + 1, column=8).font = Font(bold=True, size=12)
                ws_cot.cell(row=ultima_fila + 1, column=7).fill = PatternFill(
                    start_color="4CAF50", end_color="4CAF50", fill_type="solid"
                )
                ws_cot.cell(row=ultima_fila + 1, column=8).fill = PatternFill(
                    start_color="4CAF50", end_color="4CAF50", fill_type="solid"
                )
                ws_cot.cell(row=ultima_fila + 1, column=8).number_format = '"$"#,##0.00'
                
                # Bordes
                borde = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws_cot.iter_rows(min_row=1, max_row=ultima_fila + 1,
                                            min_col=1, max_col=8):
                    for cell in row:
                        cell.border = borde
                
                # Formato de números
                for row in range(2, ultima_fila + 1):
                    ws_cot.cell(row=row, column=5).number_format = '"$"#,##0.00'
                    ws_cot.cell(row=row, column=7).number_format = '"$"#,##0.00'
                    ws_cot.cell(row=row, column=8).number_format = '"$"#,##0.00'
            
            messagebox.showinfo(
                "✅ Exportación Exitosa",
                f"Cotización exportada correctamente:\n\n"
                f"📁 {os.path.basename(ruta_archivo)}\n\n"
                f"💰 Total: ${total:,.2f}\n"
                f"📊 Conceptos: {len(self.cotizacion_actual)}"
            )
            
            # Preguntar si desea limpiar
            if messagebox.askyesno(
                "Nueva Cotización",
                "¿Desea limpiar la cotización actual para crear una nueva?"
            ):
                self.limpiar_cotizacion()
            
        except Exception as e:
            messagebox.showerror(
                "❌ Error al Exportar",
                f"Ocurrió un error al generar el archivo:\n\n{str(e)}"
            )


# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

def main():
    """Función principal que inicia la aplicación."""
    print("\n")
    print("╔" + "="*68 + "╗")
    print("║" + " "*68 + "║")
    print("║" + "  SISTEMA DE COTIZACIÓN REMAA - v2.0  ".center(68) + "║")
    print("║" + " "*68 + "║")
    print("╚" + "="*68 + "╝")
    print("\n")
    print("🚀 Iniciando aplicación...")
    print("📂 Cargando conceptos desde: conceptos_master.csv")
    print("\n")
    
    root = tk.Tk()
    app = SistemaCotizacionREMAA(root)
    root.mainloop()


if __name__ == "__main__":
    main()